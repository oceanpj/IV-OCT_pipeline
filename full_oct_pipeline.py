"""
full_oct_pipeline.py
--------------------
Simplified, consolidated OCT pipeline.

For each case discovered in DATA_DIR the pipeline:

  A) Copies  original.dcm     — from round_color/ (no recomputation)
  B) Builds  extended_mask.dcm — fills NII annotation wedge, warps to
                                 Cartesian (skipped if no NII found)
  C) Records one row in        data_summary.xlsx — BIN file identifier,
                                 all available user metadata, output references

Output layout per case:
    Analysis/<UUID>/
        original.dcm          ← copied verbatim from round_color/<UUID>.dcm
        extended_mask.dcm     ← generated from <UUID>.nii annotation (if found)

Aggregate output:
    Analysis/
        data_summary.xlsx     ← single consolidated metadata table (all cases)

Case discovery is robust: only raw.bin + hardware.inf are mandatory.
NII absence → extended_mask step is skipped and noted in the Excel.
DICOM absence → original.dcm step is skipped and noted in the Excel.
All 20 BIN cases are found regardless of whether optional files are present.
"""

import os
import re
import struct
import shutil
import datetime
import numpy as np
import nibabel as nib
import pydicom
import cv2
from scipy.ndimage import map_coordinates
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────
DATA_DIR     = "/Users/oceanpunsalan/Data/Intravascular/20_raw"
ANALYSIS_DIR = "/Users/oceanpunsalan/Data/Intravascular/Analysis"

# ── Global settings ───────────────────────────────────────────────────────────
EXPORT_DIM  = 512   # Cartesian output resolution (px × px)
THETA_STEPS = 720   # angular resolution for NII polar conversion


# ══════════════════════════════════════════════════════════════════════════════
#  Robust case discovery
# ══════════════════════════════════════════════════════════════════════════════

def _find_nii(data_dir, case_folder, entry):
    """
    Locate the NII annotation for a case using a multi-location search.

    Search order:
      1. <data_dir>/<UUID>.nii   or  <data_dir>/<UUID>.nii.gz   (standard)
      2. <case_folder>/<UUID>.nii or  <case_folder>/<UUID>.nii.gz
      3. Any *.nii / *.nii.gz file inside <case_folder>

    Returns the first match as an absolute path, or None if nothing is found.
    Not finding a NII is not an error — extended_mask will simply be skipped.
    """
    for base in (data_dir, case_folder):
        for ext in ('.nii', '.nii.gz'):
            p = os.path.join(base, entry + ext)
            if os.path.exists(p):
                return p
    try:
        for fname in sorted(os.listdir(case_folder)):
            if fname.endswith('.nii') or fname.endswith('.nii.gz'):
                return os.path.join(case_folder, fname)
    except OSError:
        pass
    return None


def _find_dicom(case_folder, entry):
    """
    Locate the original round-color DICOM for a case.

    Search order:
      1. <case_folder>/round_color/<UUID>.dcm   (standard exact match)
      2. Any *.dcm inside <case_folder>/round_color/
      3. Any *.dcm directly inside <case_folder>  (files only, not sub-dirs)

    Returns the first match as an absolute path, or None if nothing is found.
    """
    # 1. Standard location
    p = os.path.join(case_folder, 'round_color', entry + '.dcm')
    if os.path.exists(p):
        return p

    # 2. Any DCM inside round_color/
    rc = os.path.join(case_folder, 'round_color')
    if os.path.isdir(rc):
        for fname in sorted(os.listdir(rc)):
            if fname.lower().endswith('.dcm'):
                return os.path.join(rc, fname)

    # 3. Any DCM at the top level of the case folder
    try:
        for fname in sorted(os.listdir(case_folder)):
            fpath = os.path.join(case_folder, fname)
            if os.path.isfile(fpath) and fname.lower().endswith('.dcm'):
                return fpath
    except OSError:
        pass

    return None


def find_cases(data_dir):
    """
    Discover all valid cases in data_dir.

    A case is any UUID-named subdirectory that has round_color/ (for DCM).
    The raw.bin and hardware.inf are optional — if present, metadata is extracted.
    NII and DICOM paths are located via flexible search.

    Returns
    -------
    list of (uuid, case_folder, nii_path | None, dcm_path | None)
    """
    cases = []

    try:
        entries = sorted(os.listdir(data_dir))
    except FileNotFoundError:
        print(f"[ERROR] Data directory not found: {data_dir}")
        return []
    except PermissionError:
        print(f"[ERROR] Permission denied reading: {data_dir}")
        return []

    for entry in entries:
        case_folder = os.path.join(data_dir, entry)

        if not os.path.isdir(case_folder):
            continue  # skip regular files

        # Skip if no round_color folder (need this for DCM)
        if not os.path.isdir(os.path.join(case_folder, 'round_color')):
            continue

        nii_path = _find_nii(data_dir, case_folder, entry)
        dcm_path = _find_dicom(case_folder, entry)

        if nii_path is None:
            print(f"  [warn] {entry}: NII not found — extended_mask will be skipped")
        if dcm_path is None:
            print(f"  [warn] {entry}: round-color DICOM not found — DCM copy will be skipped")

        cases.append((entry, case_folder, nii_path, dcm_path))

    return cases


# ══════════════════════════════════════════════════════════════════════════════
#  User metadata extraction
# ══════════════════════════════════════════════════════════════════════════════

def read_hardware_inf_full(path):
    """
    Parse hardware.inf.

    Returns
    -------
    params    : dict — numeric acquisition parameters (FrameLines, FrameRate,
                       PbackLength, PbackSpeed, RecordPoints, etc.)
    user_meta : dict — device version strings / identifiers present in the file:
                         Software      — build ID string (e.g. "D39933C94-041825")
                         ZHK_HV/SV/SN/PN — ZHK device hardware / software / serial / part
                         DOC_HV/SV/SN/PN — DOC device hardware / software / serial / part
                         Timestamp     — acquisition timestamp if present
    """
    params    = {}
    user_meta = {}

    try:
        with open(path, 'r', errors='replace') as f:
            lines = [l.rstrip() for l in f]
    except OSError:
        return params, user_meta

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # Device version lines — ZHK: X(HV),X(SV),X(SN),X(PN)
        dev_m = re.match(r'^(ZHK|DOC)\s*:\s*(.+)$', line)
        if dev_m:
            dev, rest = dev_m.group(1), dev_m.group(2)
            for seg in rest.split(','):
                seg = seg.strip()
                sub = re.match(r'^([^(]+)\((\w+)\)$', seg)
                if sub:
                    user_meta[f'{dev}_{sub.group(2)}'] = sub.group(1).strip()
            continue

        # Software version line — "Software: <build-id>"
        sw_m = re.match(r'^Software\s*:\s*(.+)$', line, re.IGNORECASE)
        if sw_m:
            user_meta['Software'] = sw_m.group(1).strip()
            continue

        # Timestamp line — e.g. "1970-01-01 08:00:00"
        ts_m = re.match(r'^(\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}:\d{2})\s*$', line)
        if ts_m:
            user_meta['Timestamp'] = ts_m.group(1)
            continue

        # Standard key:value acquisition parameters (may be comma-separated)
        for token in line.split(','):
            token = token.strip()
            if ':' not in token:
                continue
            key, val = token.split(':', 1)
            key = key.strip()
            val = val.strip()
            if key in ('ZHK', 'DOC', 'Software'):
                continue
            num_str = ''.join(c for c in val if c.isdigit() or c == '.')
            try:
                params[key] = float(num_str) if '.' in num_str else int(num_str)
            except ValueError:
                params[key] = val

    return params, user_meta


def read_userdata_bin(path):
    """
    Read userdata.bin → dict of raw uint32 fields (big-endian).

    Keys use the convention established in extract_data.py:
      magic    — file-type magic constant (hex string)
      field_1  through field_4 — raw uint32 values
    Returns empty dict if file is absent or unreadable.
    """
    if not os.path.exists(path):
        return {}
    try:
        with open(path, 'rb') as f:
            data = f.read()
    except OSError:
        return {}
    n = len(data) // 4
    if n == 0:
        return {}
    vals = struct.unpack(f'>{n}I', data[:n * 4])
    keys = ['magic', 'field_1', 'field_2', 'field_3', 'field_4']
    return {
        (keys[i] if i < len(keys) else f'field_{i}'): (hex(v) if i == 0 else v)
        for i, v in enumerate(vals)
    }


def read_dicom_study_meta(dcm_path):
    """
    Read study-level DICOM metadata using the official tag keyword names
    that are actually present in the round_color files.

    Returns keys: StudyDate, StudyTime, StudyInstanceUID, Modality,
                  NumberOfFrames.  Empty dict if file is absent or unreadable.
    """
    if not dcm_path or not os.path.exists(dcm_path):
        return {}
    try:
        ds = pydicom.dcmread(dcm_path, stop_before_pixels=True)
        return {
            'StudyDate':        str(getattr(ds, 'StudyDate',        '') or ''),
            'StudyTime':        str(getattr(ds, 'StudyTime',        '') or ''),
            'StudyInstanceUID': str(getattr(ds, 'StudyInstanceUID', '') or ''),
            'Modality':         str(getattr(ds, 'Modality',         '') or ''),
            'NumberOfFrames':   int(getattr(ds, 'NumberOfFrames',   0)  or 0),
        }
    except Exception as exc:
        print(f"    [warn] DICOM metadata read failed ({dcm_path}): {exc}")
        return {}


# ══════════════════════════════════════════════════════════════════════════════
#  Extended mask from NII annotation
# ══════════════════════════════════════════════════════════════════════════════

def _c2p(image, center=None, r_max=None, theta_steps=THETA_STEPS):
    """Cartesian → polar transform (NII space)."""
    h, w = image.shape
    if center is None:
        center = (h / 2.0, w / 2.0)
    if r_max is None:
        r_max = int(min(center[0], center[1], h - center[0], w - center[1]))
    rs     = np.arange(r_max)
    thetas = np.linspace(0, 2 * np.pi, theta_steps, endpoint=False)
    rg, tg = np.meshgrid(rs, thetas, indexing='ij')
    xs = center[1] + rg * np.cos(tg)
    ys = center[0] + rg * np.sin(tg)
    return map_coordinates(image, [ys, xs], order=0, mode='constant', cval=0)


def _p2c(polar, output_shape, center=None):
    """Polar → Cartesian transform (NII space)."""
    h, w = output_shape
    if center is None:
        center = (h / 2.0, w / 2.0)
    r_max, theta_steps = polar.shape
    ys, xs = np.mgrid[0:h, 0:w]
    dy    = ys - center[0]
    dx    = xs - center[1]
    r     = np.sqrt(dy**2 + dx**2).clip(0, r_max - 1)
    theta = np.arctan2(dy, dx) % (2 * np.pi)
    t_idx = theta / (2 * np.pi) * theta_steps
    return map_coordinates(polar, [r, t_idx], order=0, mode='constant', cval=0)


def _fill_annotations_polar(polar):
    """Fill thin arc annotations radially inward to produce a solid wedge."""
    filled = np.zeros_like(polar)
    labels = np.unique(polar)
    labels = labels[labels != 0]
    label_radii = {}
    for label in labels:
        rows, _ = np.where(polar == label)
        label_radii[label] = np.median(rows) if len(rows) > 0 else 0
    for label in sorted(labels, key=lambda l: label_radii[l]):
        mask = (polar == label)
        for col in range(polar.shape[1]):
            hits = np.where(mask[:, col])[0]
            if hits.size == 0:
                continue
            filled[0: hits[-1] + 1, col] = label
    return filled


def _orient(arr):
    """Orientation chain matching main.py: reflect + flip."""
    reflected = np.flipud(np.fliplr(arr)).T
    out       = np.flipud(reflected)
    return np.fliplr(out)


def compute_extended_mask(nii_path, n_frames, output_size):
    """
    Load NII annotation, fill polar wedge per slice, warp to Cartesian.

    Returns
    -------
    (n_frames, output_size, output_size) uint8 volume
    """
    nii_data     = nib.load(nii_path).get_fdata()
    n_nii_slices = nii_data.shape[-1]
    mask_vol     = np.zeros((n_frames, output_size, output_size), dtype=np.uint8)

    for z in range(min(n_frames, n_nii_slices)):
        sl = nii_data[..., z]
        if np.all(sl == 0):
            continue
        polar        = _c2p(sl)
        filled_polar = _fill_annotations_polar(polar)
        filled_cart  = _p2c(filled_polar, sl.shape)
        oriented     = _orient(filled_cart)
        resized = cv2.resize(
            oriented.astype(np.float32),
            (output_size, output_size),
            interpolation=cv2.INTER_NEAREST
        )
        mask_vol[z] = np.clip(resized, 0, 255).astype(np.uint8)

    return mask_vol


# ══════════════════════════════════════════════════════════════════════════════
#  DICOM writer
# ══════════════════════════════════════════════════════════════════════════════

def save_dicom(volume, path):
    """Save (n_frames, H, W) uint8 volume as a multi-frame DICOM."""
    now       = datetime.datetime.now()
    file_meta = pydicom.Dataset()
    file_meta.MediaStorageSOPClassUID    = pydicom.uid.CTImageStorage
    file_meta.MediaStorageSOPInstanceUID = pydicom.uid.generate_uid()
    file_meta.TransferSyntaxUID          = pydicom.uid.ExplicitVRLittleEndian

    ds = pydicom.Dataset()
    ds.file_meta        = file_meta
    ds.is_implicit_VR   = False
    ds.is_little_endian = True
    ds.SOPClassUID       = pydicom.uid.CTImageStorage
    ds.SOPInstanceUID    = file_meta.MediaStorageSOPInstanceUID
    ds.StudyInstanceUID  = pydicom.uid.generate_uid()
    ds.SeriesInstanceUID = pydicom.uid.generate_uid()
    ds.PatientName       = 'Anonymous'
    ds.PatientID         = 'ANON001'
    ds.PatientBirthDate  = ''
    ds.PatientSex        = ''
    ds.StudyDate         = now.strftime('%Y%m%d')
    ds.StudyTime         = now.strftime('%H%M%S')
    ds.Modality          = 'OT'
    ds.InstanceNumber    = '1'
    ds.NumberOfFrames    = volume.shape[0]
    ds.Rows              = volume.shape[1]
    ds.Columns           = volume.shape[2]
    ds.SamplesPerPixel           = 1
    ds.PhotometricInterpretation = 'MONOCHROME2'
    ds.BitsAllocated     = 8
    ds.BitsStored        = 8
    ds.HighBit           = 7
    ds.PixelRepresentation = 0
    ds.PixelData         = volume.astype(np.uint8).tobytes()
    pydicom.dcmwrite(path, ds, write_like_original=False)
    print(f"    Saved → {path}")


# ══════════════════════════════════════════════════════════════════════════════
#  Envelope.bin reader — generates output.dcm
# ══════════════════════════════════════════════════════════════════════════════

ENV_HEADER_BYTES = 56
ENV_MAGIC        = bytes.fromhex('a0b0cfd6')


def read_envelope_bin(path):
    """
    Read envelope.bin → (n_frames, n_alines, render_pts) uint8 OCT volume.

    Returns None if file not found or unreadable.
    """
    if not os.path.exists(path):
        return None
    try:
        with open(path, 'rb') as f:
            raw_hdr = f.read(ENV_HEADER_BYTES)
            if len(raw_hdr) < ENV_HEADER_BYTES:
                print(f"    [warn] envelope.bin header too short")
                return None
            if raw_hdr[:4] != ENV_MAGIC:
                print(f"    [warn] envelope.bin: unexpected magic bytes")
                return None

            # Parse header (13 uint32 fields after 4-byte magic)
            hdr_ints = struct.unpack('>13I', raw_hdr[4:56])
            n_alines    = hdr_ints[0]
            render_pts  = hdr_ints[1]
            n_frames    = hdr_ints[2]

            data = np.frombuffer(f.read(), dtype=np.uint8)

        # Validate expected size
        expected = n_frames * n_alines * render_pts
        if len(data) < expected:
            print(f"    [warn] envelope.bin: truncated (got {len(data)}, expected {expected})")
            return None

        volume = data[:expected].reshape(n_frames, n_alines, render_pts)
        print(f"    envelope.bin loaded: shape={volume.shape}")
        return volume

    except Exception as exc:
        print(f"    [warn] envelope.bin read failed: {exc}")
        return None


# ══════════════════════════════════════════════════════════════════════════════
#  Excel export — data_summary.xlsx
# ══════════════════════════════════════════════════════════════════════════════

# Column definitions: (header label, data-extractor callable)
# Each callable receives the full case record dict.
_COLUMNS = [
    # ── Case Identification (cols 1–4) ────────────────────────────────────────
    ('Case UUID',                lambda d: d['case_id']),
    ('raw.bin Source File',      lambda d: d['source_bin_file']),
    ('NII Annotation File',      lambda d: d.get('nii_path', '') or ''),
    ('Processing Status',        lambda d: d.get('status', '')),
    # ── Output File References (cols 5–6) ────────────────────────────────────
    ('extended_mask.dcm',        lambda d: d.get('extended_mask_path', '') or ''),
    ('original.dcm',             lambda d: d.get('original_dcm_path',  '') or ''),
    # ── Computed Parameters (cols 7–10) ─────────────────────────────────────
    ('Frame Count',              lambda d: d.get('n_frames',  '')),
    ('A-Line Count',             lambda d: d.get('n_alines',  '')),
    ('Sample Count',             lambda d: d.get('n_samples', '')),
    ('Depth Bins',               lambda d: d.get('depth',     '')),
    # ── Acquisition Parameters — hardware.inf (cols 11–25) ────────────────────
    ('FrameLines',               lambda d: d['params'].get('FrameLines',    '')),
    ('FrameRate',                lambda d: d['params'].get('FrameRate',     '')),
    ('PbackLength (mm)',         lambda d: d['params'].get('PbackLength',   '')),
    ('PbackSpeed (mm/s)',        lambda d: d['params'].get('PbackSpeed',    '')),
    ('RecordPoints',             lambda d: d['params'].get('RecordPoints',  '')),
    ('RenderPoints',             lambda d: d['params'].get('RenderPoints',  '')),
    ('LinePoints',               lambda d: d['params'].get('LinePoints',    '')),
    ('RangeInAir',               lambda d: d['params'].get('RangeInAir',   '')),
    ('ShowRange',                lambda d: d['params'].get('ShowRange',     '')),
    ('RotationSpeed',            lambda d: d['params'].get('RotationSpeed', '')),
    ('DataType',                 lambda d: d['params'].get('DataType',      '')),
    ('OverPulse',                lambda d: d['params'].get('OverPulse',     '')),
    ('RedLight',                 lambda d: d['params'].get('RedLight',      '')),
    ('Brightness',               lambda d: d['params'].get('Brightness',    '')),
    ('Contrast',                 lambda d: d['params'].get('Contrast',      '')),
    # ── Device & Software Metadata — hardware.inf (cols 26–35) ───────────────
    ('Software',                 lambda d: d['hw_user_meta'].get('Software',   '')),
    ('ZHK_HV',                   lambda d: d['hw_user_meta'].get('ZHK_HV',    '')),
    ('ZHK_SV',                   lambda d: d['hw_user_meta'].get('ZHK_SV',    '')),
    ('ZHK_SN',                   lambda d: d['hw_user_meta'].get('ZHK_SN',    '')),
    ('ZHK_PN',                   lambda d: d['hw_user_meta'].get('ZHK_PN',    '')),
    ('DOC_HV',                   lambda d: d['hw_user_meta'].get('DOC_HV',    '')),
    ('DOC_SV',                   lambda d: d['hw_user_meta'].get('DOC_SV',    '')),
    ('DOC_SN',                   lambda d: d['hw_user_meta'].get('DOC_SN',    '')),
    ('DOC_PN',                   lambda d: d['hw_user_meta'].get('DOC_PN',    '')),
    ('Timestamp',                lambda d: d['hw_user_meta'].get('Timestamp',  '')),
    # ── Userdata Binary Fields — userdata.bin (cols 36–40) ───────────────────
    ('magic',                    lambda d: d['userdata'].get('magic',   '')),
    ('field_1',                  lambda d: d['userdata'].get('field_1', '')),
    ('field_2',                  lambda d: d['userdata'].get('field_2', '')),
    ('field_3',                  lambda d: d['userdata'].get('field_3', '')),
    ('field_4',                  lambda d: d['userdata'].get('field_4', '')),
    # ── DICOM Study Metadata — round_color DICOM (cols 41–45) ────────────────
    ('StudyDate',                lambda d: d['dicom_meta'].get('StudyDate',        '')),
    ('StudyTime',                lambda d: d['dicom_meta'].get('StudyTime',        '')),
    ('StudyInstanceUID',         lambda d: d['dicom_meta'].get('StudyInstanceUID', '')),
    ('Modality',                 lambda d: d['dicom_meta'].get('Modality',         '')),
    ('DICOM NumberOfFrames',     lambda d: d['dicom_meta'].get('NumberOfFrames',   '')),
]

def export_userdata_excel(all_case_data, out_path):
    """
    Write (or overwrite) data_summary.xlsx.

    Simple format: headers in row 1, data below, narrow columns.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Cases'

    # Simple header row
    for c, (label, _) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(1, c)
        cell.value = label
        cell.font = Font(bold=True)

    # Data rows
    for r, d in enumerate(all_case_data, start=2):
        for c, (_, fn) in enumerate(_COLUMNS, start=1):
            try:
                v = fn(d)
            except Exception:
                v = ''
            ws.cell(r, c).value = v if v is not None else ''

    # Narrow column widths - fixed small sizes
    for c in range(1, len(_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12

    # Wider for specific columns that need it
    ws.column_dimensions['A'].width = 40  # Case UUID
    ws.column_dimensions['B'].width = 70  # Source file path
    ws.column_dimensions['C'].width = 50  # NII path
    ws.column_dimensions['D'].width = 30  # Status

    ws.freeze_panes = 'A2'
    wb.save(out_path)
    print(f"\n  data_summary.xlsx saved → {out_path}")


# ══════════════════════════════════════════════════════════════════════════════
#  Per-case pipeline
# ══════════════════════════════════════════════════════════════════════════════

def process_case(case_id, case_folder, nii_path, dcm_path):
    """
    Run the simplified pipeline for one case.

    Steps
    -----
    A) Copy original.dcm from round_color source (no recomputation).
    B) Generate extended_mask.dcm from NII annotation (skipped if no NII).
    C) Collect all user metadata.

    Returns
    -------
    dict — complete record for this case, ready for the Excel row.
    """
    out_dir = os.path.join(ANALYSIS_DIR, case_id)
    os.makedirs(out_dir, exist_ok=True)

    print(f"\n{'=' * 64}")
    print(f"  Case : {case_id}")
    print(f"{'=' * 64}")

    # ── Read hardware parameters + user metadata ──────────────────────────────
    hw_path = os.path.join(case_folder, 'hardware.inf')
    if os.path.exists(hw_path):
        params, hw_user_meta = read_hardware_inf_full(hw_path)
    else:
        params = {}
        hw_user_meta = {}
        print(f"  [info] hardware.inf not found — using default/empty values")

    n_alines    = int(params.get('FrameLines',   1000))
    n_samples   = int(params.get('RecordPoints', 1280))
    frame_rate  = int(params.get('FrameRate',    100))
    pback_len   = float(params.get('PbackLength', 90))
    pback_speed = float(params.get('PbackSpeed',  36))
    n_frames    = int(round(pback_len / pback_speed * frame_rate))
    depth       = n_samples // 2

    print(f"  n_frames={n_frames}  n_alines={n_alines}  n_samples={n_samples}")

    userdata_path = os.path.join(case_folder, 'userdata.bin')
    userdata      = read_userdata_bin(userdata_path)
    dicom_meta    = read_dicom_study_meta(dcm_path)

    status_parts      = []
    original_dcm_out  = None
    extended_mask_out = None

    # ── A) Copy DICOM from round_color (keep original filename) ──────────────
    if dcm_path:
        dcm_filename = os.path.basename(dcm_path)
        original_dcm_out = os.path.join(out_dir, dcm_filename)
        shutil.copy2(dcm_path, original_dcm_out)
        print(f"  [A] {dcm_filename}  ← {dcm_path}")
        status_parts.append(dcm_filename)
    else:
        print(f"  [A] DICOM SKIPPED — no source DICOM found")
        status_parts.append('SKIP:DICOM')

    # ── B) Extended mask from NII ─────────────────────────────────────────────
    if nii_path:
        print(f"  [B] Building extended_mask.dcm from {nii_path} ...")
        mask_vol = compute_extended_mask(nii_path, n_frames, EXPORT_DIM)
        extended_mask_out = os.path.join(out_dir, 'extended_mask.dcm')
        save_dicom(mask_vol, extended_mask_out)
        status_parts.append('extended_mask.dcm')
    else:
        print(f"  [B] extended_mask.dcm SKIPPED — no NII annotation found")
        status_parts.append('SKIP:extended_mask.dcm')

    status = ' | '.join(status_parts)

    # Collect case data for aggregate Excel
    raw_bin_path = os.path.join(case_folder, 'raw.bin')
    case_data = {
        # Identification
        'case_id':             case_id,
        'source_bin_file':     raw_bin_path if os.path.exists(raw_bin_path) else '',
        'nii_path':            nii_path   or '',
        'status':              status,
        # Output paths
        'extended_mask_path':  extended_mask_out or '',
        'original_dcm_path':   original_dcm_out  or '',
        # Source paths for metadata
        'hw_path':             hw_path if os.path.exists(hw_path) else '',
        'userdata_path':       userdata_path if os.path.exists(userdata_path) else '',
        # Acquisition parameters
        'params':              params,
        'n_frames':            n_frames,
        'n_alines':            n_alines,
        'n_samples':           n_samples,
        'depth':               depth,
        # User metadata
        'hw_user_meta':        hw_user_meta,
        'userdata':            userdata,
        'dicom_meta':          dicom_meta,
    }

    print(f"  ✓ Case complete  →  {out_dir}")
    return case_data


# ══════════════════════════════════════════════════════════════════════════════
#  Entry point
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    os.makedirs(ANALYSIS_DIR, exist_ok=True)

    print(f"Scanning: {DATA_DIR}")
    cases = find_cases(DATA_DIR)
    print(f"Found {len(cases)} case(s).\n")

    all_case_data = []

    for case_id, case_folder, nii_path, dcm_path in cases:
        try:
            result = process_case(case_id, case_folder, nii_path, dcm_path)
            if result is not None:
                all_case_data.append(result)
        except Exception as exc:
            import traceback
            print(f"\n  [ERROR] {case_id}: {exc}")
            traceback.print_exc()

    # ── Write consolidated data_summary ────────────────────────────────────────
    if all_case_data:
        excel_path = os.path.join(ANALYSIS_DIR, 'data_summary.xlsx')
        export_userdata_excel(all_case_data, excel_path)
    else:
        print("\n[warn] No cases processed — data_summary.xlsx not written.")

    total   = len(cases)
    success = len(all_case_data)
    print(f"\n{'=' * 64}")
    print(f"  Complete: {success}/{total} cases processed.")
    if success < total:
        print(f"  {total - success} case(s) encountered errors — see output above.")
    print(f"{'=' * 64}")
"""
extract_data.py
---------------
Reads all binary data files from an intravascular OCT+NIRS pullback.

Files
-----
hardware.inf     → acquisition parameters (text key-value)
raw.bin          → raw interferometric signal  (uint16, 250 × 1000 × 1280)
envelope.bin     → processed envelope image    (uint8,  250 × 1000 × 1024)
plague.bin       → plaque detection points     (aline, depth double pairs)
stentpoint.bin   → stent strut coordinates     (double triples)
userdata.bin     → tiny metadata blob
detail2.nirs     → NIRS chemogram data         (uint8, 155 frames)
measurement1.bin → empty / reserved
"""

import os
import struct
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")

def p(name):
    return os.path.join(DATA_DIR, name)


# ══════════════════════════════════════════════════════════════════════════════
#  hardware.inf — acquisition parameters
# ══════════════════════════════════════════════════════════════════════════════

def read_hardware_inf(path):
    params = {}
    with open(path, 'r') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            for token in line.split(','):
                token = token.strip()
                if ':' in token:
                    key, val = token.split(':', 1)
                    val = val.strip()
                    num = ''.join(c for c in val if c.isdigit() or c == '.')
                    try:
                        params[key.strip()] = float(num) if '.' in num else int(num)
                    except ValueError:
                        params[key.strip()] = val
    return params


# ══════════════════════════════════════════════════════════════════════════════
#  raw.bin — raw interferometric signal
# ══════════════════════════════════════════════════════════════════════════════

RAW_HEADER_BYTES = 28
RAW_MAGIC        = bytes.fromhex('a0b0cfd6')

def read_raw(path, n_frames=250, n_samples=1280):
    with open(path, 'rb') as f:
        hdr = f.read(RAW_HEADER_BYTES)
        assert hdr[:4] == RAW_MAGIC, "raw.bin: unexpected magic bytes"
        data = np.frombuffer(f.read(), dtype='>u2')

    # Derive n_alines from actual file size
    n_alines = len(data) // (n_frames * n_samples)
    total    = n_frames * n_alines * n_samples
    return data[:total].reshape(n_frames, n_alines, n_samples)


# ══════════════════════════════════════════════════════════════════════════════
#  envelope.bin — processed envelope (log-compressed, uint8)
# ══════════════════════════════════════════════════════════════════════════════
# FIX: Header is 56 bytes total. After the 4-byte magic, 52 bytes remain
#      (56 - 4 = 52). 52 / 4 = 13 uint32 fields — NOT 14.
#      The original code tried to unpack 14 × 4 = 56 bytes from a 52-byte
#      slice, which caused the struct.error.

ENV_HEADER_BYTES = 56
ENV_MAGIC        = bytes.fromhex('a0b0cfd6')

def read_envelope(path, n_frames=250, render_pts=1024):
    with open(path, 'rb') as f:
        raw_hdr = f.read(ENV_HEADER_BYTES)
        assert raw_hdr[:4] == ENV_MAGIC, "envelope.bin: unexpected magic bytes"

        hdr_ints = struct.unpack('>13I', raw_hdr[4:56])
        hdr = {
            'n_alines':    hdr_ints[0],
            'render_pts':  hdr_ints[1],
            'n_frames':    hdr_ints[2],
            'pullback_mm': hdr_ints[3],
            'frame_rate':  hdr_ints[4],
        }

        data = np.frombuffer(f.read(), dtype=np.uint8)

    # Derive actual n_alines from file size rather than hardcoding
    n_alines = len(data) // (n_frames * render_pts)
    hdr['actual_n_alines'] = n_alines
    return data.reshape(n_frames, n_alines, render_pts), hdr


# ══════════════════════════════════════════════════════════════════════════════
#  plague.bin — plaque detection points
# ══════════════════════════════════════════════════════════════════════════════

def read_plague(path):
    with open(path, 'rb') as f:
        hdr = struct.unpack('>4I', f.read(16))
        n_points = hdr[0]
        raw = f.read(n_points * 16)

    points = []
    for i in range(n_points):
        aline, depth = struct.unpack('>dd', raw[i*16 : i*16+16])
        if not (aline == 0.0 and depth == 0.0):
            points.append({'aline': aline, 'depth': depth})
    return points


# ══════════════════════════════════════════════════════════════════════════════
#  stentpoint.bin — stent strut coordinates
# ══════════════════════════════════════════════════════════════════════════════

def read_stentpoint(path):
    with open(path, 'rb') as f:
        f.read(16)  # skip header
        raw = f.read()

    n_doubles = len(raw) // 8
    coords = np.frombuffer(raw[:n_doubles * 8], dtype='>f8')
    n_entries = len(coords) // 3
    return coords[:n_entries * 3].reshape(n_entries, 3)


# ══════════════════════════════════════════════════════════════════════════════
#  userdata.bin — metadata blob
# ══════════════════════════════════════════════════════════════════════════════

def read_userdata(path):
    with open(path, 'rb') as f:
        data = f.read()
    n_fields = len(data) // 4
    vals = struct.unpack(f'>{n_fields}I', data[:n_fields * 4])
    keys = ['magic', 'field_1', 'field_2', 'field_3', 'field_4']
    result = {}
    for i, v in enumerate(vals):
        key = keys[i] if i < len(keys) else f'field_{i}'
        result[key] = hex(v) if i == 0 else v
    return result


# ══════════════════════════════════════════════════════════════════════════════
#  detail2.nirs — NIRS chemogram data
# ══════════════════════════════════════════════════════════════════════════════

NIRS_HEADER_BYTES = 16

def read_nirs(path):
    with open(path, 'rb') as f:
        hdr_raw = f.read(NIRS_HEADER_BYTES)
        data    = np.frombuffer(f.read(), dtype=np.uint8)

    hdr_vals = struct.unpack('>4I', hdr_raw)
    hdr = {
        'n_frames_a':          hdr_vals[1],
        'n_frames_b':          hdr_vals[2],
        'n_frames_c':          hdr_vals[3],
        'bytes_per_frame':     240000,
    }

    n_frames  = hdr['n_frames_b']
    per_frame = len(data) // n_frames
    hdr['actual_bytes_per_frame'] = per_frame

    return data.reshape(n_frames, per_frame), hdr


# ══════════════════════════════════════════════════════════════════════════════
#  Excel export
# ══════════════════════════════════════════════════════════════════════════════

HEADER_FILL  = PatternFill('solid', start_color='1F4E79')
HEADER_FONT  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
SUBHDR_FILL  = PatternFill('solid', start_color='2E75B6')
SUBHDR_FONT  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
CELL_FONT    = Font(name='Arial', size=10)
ALT_FILL     = PatternFill('solid', start_color='DCE6F1')
CENTER       = Alignment(horizontal='center', vertical='center')
LEFT         = Alignment(horizontal='left',   vertical='center')
THIN         = Side(style='thin', color='BFBFBF')
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def _hdr(cell, value):
    cell.value     = value
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = CENTER
    cell.border    = BORDER

def _subhdr(cell, value):
    cell.value     = value
    cell.font      = SUBHDR_FONT
    cell.fill      = SUBHDR_FILL
    cell.alignment = CENTER
    cell.border    = BORDER

def _cell(cell, value, alt=False):
    cell.value     = value
    cell.font      = CELL_FONT
    cell.alignment = LEFT
    cell.border    = BORDER
    if alt:
        cell.fill  = ALT_FILL

def _autowidth(ws, extra=4):
    for col in ws.columns:
        length = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(length + extra, 60)


def export_to_excel(data, out_path):
    wb = Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Summary'
    ws.row_dimensions[1].height = 22

    _hdr(ws['A1'], 'Section')
    _hdr(ws['B1'], 'Key')
    _hdr(ws['C1'], 'Value')

    rows = []

    # Acquisition params
    for k, v in data['params'].items():
        rows.append(('Acquisition Params', k, v))

    # Array shapes
    rows.append(('Array Shapes', 'envelope', str(data['envelope'].shape)))
    rows.append(('Array Shapes', 'raw',      str(data['raw'].shape)))
    rows.append(('Array Shapes', 'nirs', str(data['nirs'].shape) if data['nirs'].ndim > 1 else 'none (not present)'))
    rows.append(('Array Shapes', 'stent_pts', str(data['stentpoints'].shape) if len(data['stentpoints']) > 0 else 'none (not present)'))

    # Userdata
    for k, v in data['userdata'].items():
        rows.append(('Userdata', k, v))

    for i, (sec, key, val) in enumerate(rows):
        r = i + 2
        alt = i % 2 == 1
        _cell(ws.cell(r, 1), sec, alt)
        _cell(ws.cell(r, 2), key, alt)
        _cell(ws.cell(r, 3), str(val), alt)

    _autowidth(ws)

    # ── Sheet 2: Plaque Points ────────────────────────────────────────────────
    ws2 = wb.create_sheet('Plaque Points')
    ws2.row_dimensions[1].height = 22
    _hdr(ws2['A1'], '#')
    _hdr(ws2['B1'], 'A-Line')
    _hdr(ws2['C1'], 'Depth')

    for i, pt in enumerate(data['plague']):
        r = i + 2
        alt = i % 2 == 1
        _cell(ws2.cell(r, 1), i + 1, alt)
        _cell(ws2.cell(r, 2), round(pt['aline'], 4), alt)
        _cell(ws2.cell(r, 3), round(pt['depth'],  4), alt)
        for c in (1, 2, 3):
            ws2.cell(r, c).alignment = CENTER

    _autowidth(ws2)

    # ── Sheet 3: Stent Points ─────────────────────────────────────────────────
    ws3 = wb.create_sheet('Stent Points')
    ws3.row_dimensions[1].height = 22
    _hdr(ws3['A1'], '#')
    _hdr(ws3['B1'], 'X (mm)')
    _hdr(ws3['C1'], 'Y (mm)')
    _hdr(ws3['D1'], 'Z (mm)')

    stents = data['stentpoints']
    if len(stents) == 0:
        ws3.cell(2, 1).value = 'No stent data in this pullback'
        ws3.cell(2, 1).font  = CELL_FONT
    else:
        for i, (x, y, z) in enumerate(stents):
            r = i + 2
            alt = i % 2 == 1
            _cell(ws3.cell(r, 1), i + 1, alt)
            _cell(ws3.cell(r, 2), round(float(x), 6), alt)
            _cell(ws3.cell(r, 3), round(float(y), 6), alt)
            _cell(ws3.cell(r, 4), round(float(z), 6), alt)
            for c in range(1, 5):
                ws3.cell(r, c).alignment = CENTER

    _autowidth(ws3)

    wb.save(out_path)
    print(f"  Excel saved → {out_path}")


# ══════════════════════════════════════════════════════════════════════════════
#  Main
# ══════════════════════════════════════════════════════════════════════════════

def load_all(data_dir=None):
    if data_dir is None:
        data_dir = DATA_DIR

    def fp(name):
        return os.path.join(data_dir, name)

    results = {}

    print("─" * 60)
    print("  Loading hardware.inf ...")
    results['params'] = read_hardware_inf(fp('hardware.inf'))
    for k, v in results['params'].items():
        print(f"    {k}: {v}")

    print("\n  Loading envelope.bin ...")
    results['envelope'], env_hdr = read_envelope(fp('envelope.bin'))
    print(f"    shape: {results['envelope'].shape}  dtype: {results['envelope'].dtype}")
    print(f"    header: {env_hdr}")

    print("\n  Loading raw.bin ...")
    results['raw'] = read_raw(fp('raw.bin'))
    print(f"    shape: {results['raw'].shape}  dtype: {results['raw'].dtype}")

    print("\n  Loading plague.bin (plaque points) ...")
    results['plague'] = read_plague(fp('plague.bin'))
    print(f"    {len(results['plague'])} plaque points detected")
    for pt in results['plague']:
        print(f"    aline={pt['aline']:.0f}  depth={pt['depth']:.0f}")

    print("\n  Loading stentpoint.bin ...")
    if os.path.exists(fp('stentpoint.bin')):
        results['stentpoints'] = read_stentpoint(fp('stentpoint.bin'))
        print(f"    shape: {results['stentpoints'].shape}")
    else:
        results['stentpoints'] = np.empty((0, 3))
        print("    not found — skipping (no stent data in this pullback)")

    print("\n  Loading userdata.bin ...")
    results['userdata'] = read_userdata(fp('userdata.bin'))
    print(f"    {results['userdata']}")

    print("\n  Loading detail2.nirs ...")
    if os.path.exists(fp('detail2.nirs')):
        results['nirs'], nirs_hdr = read_nirs(fp('detail2.nirs'))
        print(f"    shape: {results['nirs'].shape}  dtype: {results['nirs'].dtype}")
        print(f"    header: {nirs_hdr}")
    else:
        results['nirs'] = np.empty((0,), dtype=np.uint8)
        print("    not found — skipping (no NIRS data in this pullback)")

    print("─" * 60)
    print("  All files loaded successfully.")
    return results


if __name__ == '__main__':
    DATA_DIR = r"/Users/oceanpunsalan/Library/Mobile Documents/com~apple~CloudDocs/Data/Intravascular/IntraVascular/B36087B8-BD7C-4853-8156-99A2733E125C"

    data = load_all(DATA_DIR)

    # ── Export to Excel ───────────────────────────────────────────────────────
    excel_out = os.path.join(DATA_DIR, "pullback_summary.xlsx")
    export_to_excel(data, excel_out)

    # ── Quick access ──────────────────────────────────────────────────────────
    envelope   = data['envelope']     # (250, 1000, 1024) uint8
    raw        = data['raw']          # (250, 1000, 1280) uint16
    nirs       = data['nirs']         # (155, N)          uint8
    plaque_pts = data['plague']       # list of {aline, depth}
    stent_pts  = data['stentpoints']  # (N, 3) float64
    params     = data['params']       # dict from hardware.inf